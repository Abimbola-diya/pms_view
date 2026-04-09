#!/usr/bin/env python3
"""
Comprehensive Data Extraction from ALL PMS Source Files
Extracts every single detail from XLSX, CSV, and available data
"""

import os
import csv
import json
from datetime import datetime
from collections import defaultdict
import re

DATA_DIR = "/home/abimbola/Desktop/PMS_visualization/PMS data"
OUTPUT_DIR = "/home/abimbola/Desktop/PMS_visualization/output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ================================================================
# EXTRACTION FUNCTIONS
# ================================================================

def extract_shipping_data():
    """Extract ALL shipping position data from CSV files"""
    shipping_data = []
    
    files = [f for f in os.listdir(DATA_DIR) if 'daily_shipping_position' in f.lower() and f.endswith('.csv')]
    
    print(f"  Found {len(files)} shipping files")
    
    for file in files:
        filepath = os.path.join(DATA_DIR, file)
        try:
            # Extract date from filename
            date_match = re.search(r'Sun Apr (\d+) (\d+)', file)
            file_date = f"2026-04-{int(date_match.group(1)):02d}" if date_match else "2026-04-05"
            
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get('Vessel Name', '').strip().upper() != 'VACANT':
                        # Clean data
                        vessel_name = row.get('Vessel Name', '').strip()
                        imo = row.get('IMO Number', '').strip()
                        length = row.get('Length(M)', '').strip()
                        berth = row.get('Berth', '').strip()
                        commodity = row.get('Comm', '').strip()
                        agent = row.get('Agent', '').strip()
                        berth_date = row.get('Berth Date', '').strip()
                        etd = row.get('ETD', '').strip()
                        
                        shipping_data.append({
                            'vessel_name': vessel_name,
                            'imo_number': imo,
                            'length_m': float(length) if length else None,
                            'berth_location': berth,
                            'commodity': commodity,
                            'agent': agent,
                            'berth_date': berth_date,
                            'etd': etd,
                            'data_file_date': file_date,
                            'rotation': row.get('Rotation', '').strip()
                        })
        except Exception as e:
            print(f"    Error in {file}: {e}")
    
    return shipping_data


def extract_iea_data():
    """Extract ALL IEA energy statistics"""
    iea_data = defaultdict(list)
    
    files = [f for f in os.listdir(DATA_DIR) if f.startswith('International Energy Agency') and f.endswith('.csv')]
    
    print(f"  Found {len(files)} IEA files")
    
    for file in files:
        filepath = os.path.join(DATA_DIR, file)
        file_type = file.replace('International Energy Agency - ', '').replace('.csv', '')
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                # Skip BOM if present
                content = f.read()
                if content.startswith('\ufeff'):
                    content = content[1:]
                
                lines = content.split('\n')
                reader = csv.DictReader(lines)
                
                for row in reader:
                    if not row or not any(row.values()):
                        continue
                    
                    iea_data[file_type].append({
                        'year': row.get('Year'),
                        'value': row.get(list(row.keys())[1], ''),  # Get 2nd column value
                        'unit': row.get('Units', ''),
                        'raw_data': row
                    })
        except Exception as e:
            print(f"    Error in {file}: {e}")
    
    return iea_data


def extract_csv_simple(filename_pattern, key_columns):
    """Generic CSV extraction"""
    data = []
    
    files = [f for f in os.listdir(DATA_DIR) if filename_pattern.lower() in f.lower() and f.endswith('.csv')]
    
    for file in files:
        filepath = os.path.join(DATA_DIR, file)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if any(row.get(k) for k in key_columns):
                        data.append(row)
        except Exception as e:
            print(f"    Error in {file}: {e}")
    
    return data


def extract_xlsx_simple(filename_pattern):
    """Generic XLSX extraction - returns raw data"""
    all_data = []
    
    files = [f for f in os.listdir(DATA_DIR) if filename_pattern.lower() in f.lower() and f.endswith('.xlsx')]
    
    print(f"  Found {len(files)} files matching '{filename_pattern}'")
    
    for file in files:
        filepath = os.path.join(DATA_DIR, file)
        try:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = []
                    
                    # Get headers from first row
                    for cell in ws[1]:
                        headers.append(cell.value)
                    
                    # Get all data rows
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if any(row):  # Skip empty rows
                            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                            row_dict['_file'] = file
                            row_dict['_sheet'] = sheet_name
                            row_dict['_row'] = row_idx
                            all_data.append(row_dict)
            except ImportError:
                print(f"    (openpyxl not available for {file})")
        except Exception as e:
            print(f"    Error in {file}: {e}")
    
    return all_data


# ================================================================
# COMPREHENSIVE EXTRACTION
# ================================================================

def main():
    print("\n" + "="*80)
    print("COMPREHENSIVE DATA EXTRACTION - ALL FILES")
    print("="*80 + "\n")
    
    # 1. SHIPPING DATA
    print("1. Extracting SHIPPING DATA...")
    shipping = extract_shipping_data()
    print(f"   ✓ Extracted {len(shipping)} vessel records\n")
    
    # 2. IEA DATA
    print("2. Extracting IEA ENERGY STATISTICS...")
    iea = extract_iea_data()
    total_iea = sum(len(v) for v in iea.values())
    print(f"   ✓ Extracted {total_iea} records from {len(iea)} IEA files\n")
    
    # 3. PRODUCTION DATA (NUPRC, FIRS)
    print("3. Extracting PRODUCTION DATA (NUPRC, FIRS, etc)...")
    prod_nuprc = extract_xlsx_simple('NUPRC CRUDE OIL LIFTING')
    prod_firs = extract_xlsx_simple('FIRS CRUDE OIL LIFTING')
    prod_crude = extract_xlsx_simple('CRUDE OIL PRODUCTION')
    prod_dsdp = extract_xlsx_simple('DSDP')
    
    total_prod = len(prod_nuprc) + len(prod_firs) + len(prod_crude) + len(prod_dsdp)
    print(f"   ✓ NUPRC Lifting: {len(prod_nuprc)} records")
    print(f"   ✓ FIRS Lifting: {len(prod_firs)} records")
    print(f"   ✓ Crude Production: {len(prod_crude)} records")
    print(f"   ✓ DSDP: {len(prod_dsdp)} records")
    print(f"   ✓ Total production: {total_prod} records\n")
    
    # 4. RIG DISPOSITION
    print("4. Extracting RIG DISPOSITION DATA...")
    rigs = extract_xlsx_simple('RIG-DISPOSITION')
    print(f"   ✓ Extracted {len(rigs)} rig records\n")
    
    # 5. PETROLEUM PRODUCTS STOCK
    print("5. Extracting PETROLEUM PRODUCTS STOCK DATA...")
    stock = extract_csv_simple('Petroleum-Products-Stock', ['Stock', 'Days'])
    print(f"   ✓ Extracted {len(stock)} stock records\n")
    
    # 6. OIL BLOCKS
    print("6. Extracting OIL BLOCKS DATA...")
    blocks_csv = extract_csv_simple('oil.csv', ['Block', 'Field'])
    print(f"   ✓ Extracted {len(blocks_csv)} block records\n")
    
    # SAVE RAW DATA FOR INSPECTION
    print("7. Saving raw data for inspection...")
    
    with open(os.path.join(OUTPUT_DIR, 'raw_shipping.json'), 'w') as f:
        json.dump(shipping, f, indent=2, default=str)
    
    with open(os.path.join(OUTPUT_DIR, 'raw_iea.json'), 'w') as f:
        json.dump({k: v for k, v in iea.items()}, f, indent=2, default=str)
    
    with open(os.path.join(OUTPUT_DIR, 'raw_production.json'), 'w') as f:
        json.dump({
            'nuprc': prod_nuprc,
            'firs': prod_firs,
            'crude': prod_crude,
            'dsdp': prod_dsdp
        }, f, indent=2, default=str)
    
    with open(os.path.join(OUTPUT_DIR, 'raw_rigs.json'), 'w') as f:
        json.dump(rigs, f, indent=2, default=str)
    
    with open(os.path.join(OUTPUT_DIR, 'raw_stock.json'), 'w') as f:
        json.dump(stock, f, indent=2, default=str)
    
    print("   ✓ Saved raw data JSON files\n")
    
    # PRINT SAMPLE DATA FOR REVIEW
    print("\n" + "="*80)
    print("SAMPLE DATA FOR REVIEW")
    print("="*80 + "\n")
    
    if shipping:
        print("SHIPPING - Sample vessel record:")
        print(json.dumps(shipping[0], indent=2, default=str))
    
    if prod_nuprc:
        print("\n\nPRODUCTION (NUPRC) - First record:")
        print(json.dumps(prod_nuprc[0], indent=2, default=str))
    
    if rigs:
        print("\n\nRIGS - Sample rig record:")
        print(json.dumps(rigs[0], indent=2, default=str))
    
    print("\n" + "="*80)
    print("EXTRACTION SUMMARY")
    print("="*80)
    print(f"Shipping vessels: {len(shipping)}")
    print(f"IEA records: {total_iea}")
    print(f"Production records: {total_prod}")
    print(f"Rig records: {len(rigs)}")
    print(f"Stock records: {len(stock)}")
    print(f"Block records: {len(blocks_csv)}")
    print(f"\nTotal raw records: {len(shipping) + total_iea + total_prod + len(rigs) + len(stock) + len(blocks_csv)}")
    print(f"\nRaw data saved to: {OUTPUT_DIR}/")


if __name__ == '__main__':
    main()
